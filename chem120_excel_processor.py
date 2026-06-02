"""
Excel processing logic for the CHEM 120 compiler.

Main behavior:
- Find atomic-number columns such as ZA, ZAP, ZB, ZBP, ZBDP, ZO
- Fill them from matching symbol columns A, AP, B, BP, BDP, O
- Also supports wide-format headers like 1ZA -> 1A, 2ZB -> 2B, 3ZBDP -> 3BDP
- Optionally fills PN from P and BubN from Bub
- Preserves workbook formatting because it edits the uploaded Excel workbook directly
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from atomic_numbers import atomic_number_for, normalize_symbol


@dataclass
class ProcessingReport:
    sheets_processed: int = 0
    rows_scanned: int = 0
    atomic_cells_filled: int = 0
    code_cells_filled: int = 0
    unknown_symbols: list[dict[str, Any]] = field(default_factory=list)
    skipped_sheets: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheets_processed": self.sheets_processed,
            "rows_scanned": self.rows_scanned,
            "atomic_cells_filled": self.atomic_cells_filled,
            "code_cells_filled": self.code_cells_filled,
            "unknown_symbols": self.unknown_symbols,
            "skipped_sheets": self.skipped_sheets,
        }


def _clean_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    """Return header name -> 1-based column index."""
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        header = _clean_header(cell.value)
        if header:
            headers[header] = cell.column
    return headers


def _atomic_pairs(headers: dict[str, int]) -> list[tuple[str, str]]:
    """
    Build pairs of (symbol_column, atomic_number_column).

    Examples:
        A -> ZA
        AP -> ZAP
        B -> ZB
        BP -> ZBP
        BDP -> ZBDP
        O -> ZO
        1A -> 1ZA
        3BDP -> 3ZBDP
    """
    pairs: list[tuple[str, str]] = []

    for atomic_header in headers:
        symbol_header = None

        # Long format: ZA -> A, ZBP -> BP
        if atomic_header.startswith("Z") and len(atomic_header) > 1:
            candidate = atomic_header[1:]
            if candidate in headers:
                symbol_header = candidate

        # Wide format: 1ZA -> 1A, 3ZBDP -> 3BDP
        match = re.fullmatch(r"(\d+)Z(.+)", atomic_header)
        if match:
            prefix, rest = match.groups()
            candidate = f"{prefix}{rest}"
            if candidate in headers:
                symbol_header = candidate

        if symbol_header:
            pairs.append((symbol_header, atomic_header))

    # Keep output stable and readable
    preferred_order = ["A", "AP", "B", "BP", "BDP", "O"]
    def sort_key(pair: tuple[str, str]) -> tuple[int, str]:
        symbol_header, atomic_header = pair
        stripped = re.sub(r"^\d+", "", symbol_header)
        rank = preferred_order.index(stripped) if stripped in preferred_order else 99
        return rank, atomic_header

    return sorted(set(pairs), key=sort_key)


def _code_pairs(headers: dict[str, int]) -> list[tuple[str, str, str]]:
    """
    Build pairs of (source_text_column, output_number_column, code_type).

    Examples:
        P -> PN
        Bub -> BubN
        1P -> 1PN
        2Bub -> 2BubN
    """
    pairs: list[tuple[str, str, str]] = []

    for code_header in headers:
        # Phase number: PN from P, or 1PN from 1P
        phase_match = re.fullmatch(r"(\d*)PN", code_header)
        if phase_match:
            prefix = phase_match.group(1)
            source = f"{prefix}P"
            if source in headers:
                pairs.append((source, code_header, "phase"))

        # Bubble number: BubN from Bub, or 1BubN from 1Bub
        bubble_match = re.fullmatch(r"(\d*)BubN", code_header)
        if bubble_match:
            prefix = bubble_match.group(1)
            source = f"{prefix}Bub"
            if source in headers:
                pairs.append((source, code_header, "bubble"))

    return sorted(set(pairs), key=lambda x: x[1])


_PHASE_CODES = {
    "impure": 1,
    "impure phase": 1,
    "heterogeneous": 1,
    "heterogeneous mixture": 1,
    "heterogenous mixture": 1,
    "mixed phase": 1,
    "mixture": 1,
    "pure": 2,
    "pure phase": 2,
    "homogeneous": 2,
    "homogeneous mixture": 2,
    "homogenous mixture": 2,
    "pure phase/homogenous mixture": 2,
    "pure phase/homogeneous mixture": 2,
}

_BUBBLE_CODES = {
    "maybe": 0,
    "yes": 1,
    "y": 1,
    "no": 2,
    "n": 2,
}


def _code_for(value: object, code_type: str) -> int | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text or text in {"nan", "none", "n/a", "na"}:
        return None

    if code_type == "phase":
        return _PHASE_CODES.get(text)
    if code_type == "bubble":
        return _BUBBLE_CODES.get(text)
    return None


def process_workbook_bytes(
    uploaded_bytes: bytes,
    *,
    fill_atomic_numbers: bool = True,
    fill_phase_and_bubble_codes: bool = True,
    overwrite_existing: bool = True,
) -> tuple[bytes, dict[str, Any]]:
    """
    Process an uploaded .xlsx file and return (processed_xlsx_bytes, report).

    Args:
        uploaded_bytes:
            Raw bytes from Streamlit's uploaded file.
        fill_atomic_numbers:
            Fill yellow atomic-number columns such as ZA, ZB, ZO.
        fill_phase_and_bubble_codes:
            Fill PN and BubN automatically from P and Bub.
        overwrite_existing:
            If True, correct existing values too. If False, only fill blanks.
    """
    input_buffer = BytesIO(uploaded_bytes)
    workbook = load_workbook(input_buffer)

    report = ProcessingReport()

    for ws in workbook.worksheets:
        headers = _header_map(ws)
        if not headers:
            report.skipped_sheets.append(ws.title)
            continue

        atomic_pairs = _atomic_pairs(headers) if fill_atomic_numbers else []
        code_pairs = _code_pairs(headers) if fill_phase_and_bubble_codes else []

        if not atomic_pairs and not code_pairs:
            report.skipped_sheets.append(ws.title)
            continue

        report.sheets_processed += 1

        for row in range(2, ws.max_row + 1):
            report.rows_scanned += 1

            if fill_atomic_numbers:
                for symbol_header, atomic_header in atomic_pairs:
                    symbol_col = headers[symbol_header]
                    atomic_col = headers[atomic_header]

                    symbol_cell = ws.cell(row=row, column=symbol_col)
                    atomic_cell = ws.cell(row=row, column=atomic_col)

                    if not overwrite_existing and atomic_cell.value not in (None, ""):
                        continue

                    symbol = normalize_symbol(symbol_cell.value)
                    atomic_number = atomic_number_for(symbol)

                    if symbol is None:
                        # Blank symbol means blank atomic number.
                        if overwrite_existing and atomic_cell.value not in (None, ""):
                            atomic_cell.value = None
                            report.atomic_cells_filled += 1
                        continue

                    if atomic_number is None:
                        atomic_cell.value = None
                        report.unknown_symbols.append(
                            {
                                "sheet": ws.title,
                                "row": row,
                                "symbol_column": symbol_header,
                                "atomic_column": atomic_header,
                                "entered_value": symbol_cell.value,
                            }
                        )
                        continue

                    if atomic_cell.value != atomic_number:
                        atomic_cell.value = atomic_number
                        report.atomic_cells_filled += 1

            if fill_phase_and_bubble_codes:
                for source_header, code_header, code_type in code_pairs:
                    source_col = headers[source_header]
                    code_col = headers[code_header]

                    source_cell = ws.cell(row=row, column=source_col)
                    code_cell = ws.cell(row=row, column=code_col)

                    if not overwrite_existing and code_cell.value not in (None, ""):
                        continue

                    code = _code_for(source_cell.value, code_type)
                    if code is None:
                        continue

                    if code_cell.value != code:
                        code_cell.value = code
                        report.code_cells_filled += 1

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer.getvalue(), report.as_dict()
